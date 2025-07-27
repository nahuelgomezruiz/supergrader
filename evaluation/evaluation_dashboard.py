import os
import csv
import json
import asyncio
import aiohttp
import pandas as pd
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
from datetime import datetime
import re
import logging
import random
import statistics

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('evaluation_debug.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# QWERTY order for radio button option letters
QWERTY_LETTERS = "QWERTYUIOPASDFGHJKLZXCVBNM"

# Concurrent batch processing configuration
DEFAULT_CONCURRENT_BATCH_SIZE = 10  # Number of students to process concurrently


@dataclass
class RubricItem:
    """Represents a single rubric item."""
    id: str
    description: str
    points: float
    type: str  # "CHECKBOX" or "RADIO"
    options: Optional[Dict[str, str]] = None  # For RADIO type
    column_index: int = 0  # Original column index in CSV


@dataclass
class EvaluationTask:
    """Represents a single student evaluation task."""
    project_name: str
    csv_name: str
    student_id: str
    assignment_context: Dict[str, str]
    source_files: Dict[str, str]
    rubric_items: List[Dict[str, Any]]
    human_grade: Any  # StudentGrade object
    rubric_items_list: List[Any]  # List of RubricItem objects


@dataclass 
class StudentGrade:
    """Represents a student's grades."""
    submission_id: str
    name: str
    grades: Dict[str, Any]  # Maps rubric_item_id -> grade (bool for checkbox, str for radio)


@dataclass
class AIGradingResult:
    """Represents AI grading results for a student."""
    submission_id: str
    decisions: Dict[str, Any]  # Maps rubric_item_id -> GradingDecision
    

@dataclass
class EvaluationResult:
    """Represents the result of evaluating a single rubric item."""
    submission_id: str
    rubric_id: str
    decision: str
    confidence: float
    comment: str
    evidence: Dict[str, Any]


@dataclass
class PointsComparison:
    """Comparison of points between AI and human graders."""
    submission_id: str
    human_total: float
    ai_total: float  
    difference: float
    rubric_breakdown: Dict[str, Dict[str, float]]  # rubric_id -> {"human": points, "ai": points}


@dataclass
class RubricItemStats:
    """Statistics for a single rubric item."""
    rubric_id: str
    description: str
    rubric_type: str
    max_points: float
    human_avg: float
    ai_avg: float
    human_std: float
    ai_std: float
    agreement_rate: float  # For categorical items, % agreement
    sample_size: int


def add_credit_indicators_to_radio_options(options: Dict[str, Dict[str, str]]) -> Dict[str, str]:
    """Add credit indicators to radio button options based on point values."""
    if not options:
        return {}
    
    # Find the maximum point value
    max_points = 0.0
    for option_data in options.values():
        if isinstance(option_data, dict) and 'points' in option_data:
            try:
                points = float(option_data['points'])
                max_points = max(max_points, points)
            except (ValueError, TypeError):
                pass
    
    backend_options = {}
    for letter, option_data in options.items():
        if isinstance(option_data, dict) and 'text' in option_data:
            text = option_data['text']
            try:
                points = float(option_data.get('points', 0))
                
                # Add credit indicator based on point value
                if points == max_points and points > 0:
                    text += " (Full credit)"
                elif points > 0 and points < max_points:
                    text += " (Partial credit)"
                elif points == 0:
                    text += " (No credit)"
                    
                backend_options[letter] = text
            except (ValueError, TypeError):
                # Fallback if points parsing fails
                backend_options[letter] = text + " (No credit)"
        else:
            # Fallback for old format
            backend_options[letter] = str(option_data)
    
    return backend_options


def print_radio_button_json(backend_rubric_items, context=""):
    """Print radio button JSON being sent to backend for debugging."""
    radio_items = [item for item in backend_rubric_items if item.get("type") == "RADIO"]
    
    if radio_items:
        print(f"📻 {context}Sending {len(radio_items)} radio button(s) to backend:")
        for radio_item in radio_items:
            print(f"   🎯 {radio_item['id']}: {radio_item['description']} ({radio_item['points']} pts)")
            if 'options' in radio_item:
                for letter, text in radio_item['options'].items():
                    credit_type = "❓"
                    if text.endswith("(Full credit)"):
                        credit_type = "🟢"
                    elif text.endswith("(Partial credit)"):
                        credit_type = "🟡"
                    elif text.endswith("(No credit)"):
                        credit_type = "🔴"
                    
                    print(f"      {letter}: {credit_type} \"{text[:60]}{'...' if len(text) > 60 else ''}\"")


def filter_rubric_items_for_backend(rubric_items: List[RubricItem]) -> List[RubricItem]:
    """Filter out bonus point questions and other items that shouldn't be sent to backend."""
    filtered_items = []
    filtered_count = 0
    
    for rubric_item in rubric_items:
        # Filter out bonus point questions (case-insensitive)
        if rubric_item.description and '(bonus point)' in rubric_item.description.lower():
            print(f"🚫 Filtering out bonus point question: {rubric_item.id} - \"{rubric_item.description[:60]}...\"")
            filtered_count += 1
            continue
            
        # Also filter out zero-point items that are often just for human graders
        if rubric_item.type == 'CHECKBOX' and rubric_item.points == 0:
            print(f"🚫 Filtering out zero-point checkbox: {rubric_item.id} - \"{rubric_item.description[:60]}...\"")
            filtered_count += 1 
            continue
            
        filtered_items.append(rubric_item)
    
    if filtered_count > 0:
        print(f"✅ Filtered out {filtered_count} items. Sending {len(filtered_items)} items to backend.")
    
    return filtered_items


class PointsCalculator:
    """Calculates and compares points between AI and human graders."""
    
    @staticmethod
    def calculate_human_points(student_grade: StudentGrade, rubric_items: List[RubricItem]) -> Dict[str, float]:
        """Calculate points earned by human grader for each rubric item."""
        points_breakdown = {}
        
        for rubric_item in rubric_items:
            if rubric_item.id not in student_grade.grades:
                points_breakdown[rubric_item.id] = 0.0
                continue
                
            if rubric_item.type == "CHECKBOX":
                # Checkbox: full points if checked, 0 if not
                is_checked = student_grade.grades[rubric_item.id]
                points_breakdown[rubric_item.id] = rubric_item.points if is_checked else 0.0
                
            elif rubric_item.type == "RADIO":
                # Radio: points based on selected option
                selected_option = student_grade.grades[rubric_item.id]
                if selected_option and rubric_item.options and selected_option in rubric_item.options:
                    points_breakdown[rubric_item.id] = float(rubric_item.options[selected_option]['points'])
                else:
                    points_breakdown[rubric_item.id] = 0.0
                    
        return points_breakdown
    
    @staticmethod 
    def calculate_ai_points(ai_result: AIGradingResult, rubric_items: List[RubricItem]) -> Dict[str, float]:
        """Calculate points earned by AI grader for each rubric item."""
        points_breakdown = {}
        
        for rubric_item in rubric_items:
            if rubric_item.id not in ai_result.decisions:
                points_breakdown[rubric_item.id] = 0.0  
                continue
                
            decision = ai_result.decisions[rubric_item.id]
            
            if rubric_item.type == "CHECKBOX":
                # For checkbox: check the decision in the verdict
                if hasattr(decision, 'verdict') and hasattr(decision.verdict, 'decision'):
                    # RubricDecision.CHECK means points awarded, CROSS means no points
                    is_awarded = decision.verdict.decision.value == "CHECK"
                    points_breakdown[rubric_item.id] = rubric_item.points if is_awarded else 0.0
                else:
                    points_breakdown[rubric_item.id] = 0.0
                    
            elif rubric_item.type == "RADIO":
                # For radio: get points based on selected option
                if hasattr(decision, 'verdict') and hasattr(decision.verdict, 'selected_option'):
                    selected_option = decision.verdict.selected_option
                    if selected_option and rubric_item.options and selected_option in rubric_item.options:
                        points_breakdown[rubric_item.id] = float(rubric_item.options[selected_option]['points'])
                    else:
                        points_breakdown[rubric_item.id] = 0.0
                else:
                    points_breakdown[rubric_item.id] = 0.0
                    
        return points_breakdown
    
    @staticmethod
    def compare_submissions(
        human_grades: Dict[str, StudentGrade], 
        ai_results: Dict[str, AIGradingResult], 
        rubric_items: List[RubricItem]
    ) -> List[PointsComparison]:
        """Compare points between human and AI graders for all submissions."""
        comparisons = []
        
        for submission_id in human_grades.keys():
            if submission_id not in ai_results:
                continue
                
            human_grade = human_grades[submission_id]
            ai_result = ai_results[submission_id]
            
            # Calculate points for each grader
            human_points = PointsCalculator.calculate_human_points(human_grade, rubric_items)
            ai_points = PointsCalculator.calculate_ai_points(ai_result, rubric_items)
            
            # Calculate totals
            human_total = sum(human_points.values())
            ai_total = sum(ai_points.values())
            
            # Create rubric breakdown
            rubric_breakdown = {}
            for rubric_id in human_points.keys():
                rubric_breakdown[rubric_id] = {
                    "human": human_points.get(rubric_id, 0.0),
                    "ai": ai_points.get(rubric_id, 0.0)
                }
            
            comparison = PointsComparison(
                submission_id=submission_id,
                human_total=human_total,
                ai_total=ai_total,
                difference=ai_total - human_total,
                rubric_breakdown=rubric_breakdown
            )
            comparisons.append(comparison)
            
        return comparisons
    
    @staticmethod 
    def calculate_rubric_stats(
        comparisons: List[PointsComparison], 
        rubric_items: List[RubricItem]
    ) -> List[RubricItemStats]:
        """Calculate statistics for each rubric item."""
        stats = []
        
        for rubric_item in rubric_items:
            human_scores = []
            ai_scores = []
            agreements = 0
            
            for comparison in comparisons:
                if rubric_item.id in comparison.rubric_breakdown:
                    breakdown = comparison.rubric_breakdown[rubric_item.id]
                    human_scores.append(breakdown["human"])
                    ai_scores.append(breakdown["ai"])
                    
                    # For agreement, check if both graders gave same points
                    if abs(breakdown["human"] - breakdown["ai"]) < 0.01:  # Account for floating point
                        agreements += 1
            
            if len(human_scores) == 0:
                continue
                
            # Calculate statistics
            human_avg = statistics.mean(human_scores)
            ai_avg = statistics.mean(ai_scores)
            human_std = statistics.stdev(human_scores) if len(human_scores) > 1 else 0.0
            ai_std = statistics.stdev(ai_scores) if len(ai_scores) > 1 else 0.0
            agreement_rate = (agreements / len(human_scores)) * 100 if len(human_scores) > 0 else 0.0
            
            stat = RubricItemStats(
                rubric_id=rubric_item.id,
                description=rubric_item.description,
                rubric_type=rubric_item.type,
                max_points=rubric_item.points,
                human_avg=human_avg,
                ai_avg=ai_avg,
                human_std=human_std,
                ai_std=ai_std,
                agreement_rate=agreement_rate,
                sample_size=len(human_scores)
            )
            stats.append(stat)
            
        return stats


class CSVParser:
    """Parses Gradescope CSV files to extract rubric items and grades."""
    
    # File processing configuration (matching Chrome extension)
    MAX_FILE_SIZE = 1048576  # 1MB in characters
    TEST_FILE_MAX_CONTENT = 100  # characters for test files
    SUPPORTED_EXTENSIONS = ['.cpp', '.h', '.py', '.java', '.js', '.ts', '.c', '.cc', '.cxx']
    
    @staticmethod
    def is_test_file(file_path: str) -> bool:
        """Check if a file is a test file (same logic as Chrome extension)."""
        lower = file_path.lower()
        
        # Get just the filename (handle both / and \ separators)
        base_name = lower.replace('\\', '/').split('/')[-1]
        
        # Core files that are NEVER treated as test files
        if base_name == 'makefile' or base_name.startswith('readme'):
            return False
        
        # Core source files (.cpp and .h) - but check for test patterns
        if lower.endswith('.cpp') or lower.endswith('.h'):
            # Special case: unit_tests.h is important for grading, don't truncate it
            if base_name == 'unit_tests.h':
                return False  # Treat as core file (don't truncate)
            # Other test files with "test" in the name are test files
            elif 'test' in base_name:
                return True
            return False
        
        # Check if it's a supported extension
        has_valid_extension = any(lower.endswith(ext) for ext in CSVParser.SUPPORTED_EXTENSIONS)
        if not has_valid_extension:
            return True  # Non-source files are considered test files
        
        # Everything else is considered a test/auxiliary file
        return True
    
    @staticmethod
    def process_file_content(file_path: str, content: str) -> str | None:
        """Process file content with filtering and truncation (same as Chrome extension)."""
        
        # Check for binary data (basic check)
        if '\0' in content:
            print(f"      ⚠️  Skipping binary file: {file_path}")
            return None
        
        # Clean line endings
        content = content.replace('\r\n', '\n').replace('\r', '\n')
        
        # Apply file size limit
        if len(content) > CSVParser.MAX_FILE_SIZE:
            content = content[:CSVParser.MAX_FILE_SIZE] + '\n\n// [FILE TRUNCATED - TOO LARGE]'
            print(f"      📏 Truncated large file: {file_path} ({len(content)} chars)")
        
        # Apply test file truncation
        if CSVParser.is_test_file(file_path):
            if len(content) > CSVParser.TEST_FILE_MAX_CONTENT:
                content = content[:CSVParser.TEST_FILE_MAX_CONTENT] + '[TRIMMED]'
                print(f"      ✂️  Truncated test file: {file_path} (to {CSVParser.TEST_FILE_MAX_CONTENT} chars)")
        
        return content
    
    @staticmethod
    def parse_rubric_from_csv(csv_path: Path) -> Tuple[List[RubricItem], Dict[str, StudentGrade]]:
        """Parse CSV file to extract rubric items and student grades."""
        rubric_items = []
        student_grades = {}
        
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
            
            if len(rows) < 4:
                raise ValueError(f"CSV file {csv_path} doesn't have enough rows")
            
            # Parse headers (first row contains questions)
            headers = rows[0]
            
            # Get point values from last row
            point_values_row = rows[-1]
            
            # Process rubric items from headers
            radio_groups = {}  # Track radio button groups
            
            for col_idx, header in enumerate(headers):
                if col_idx < 8:  # Skip metadata columns
                    continue
                    
                if not header or header.strip() == "":
                    continue
                
                # Skip columns that are not meant to be graded
                excluded_columns = ["Adjustment", "Comments", "Grader", "Tags"]
                if header.strip() in excluded_columns:
                    continue
                    
                # Check if it's a radio button
                radio_match = re.match(r'\[RADIO (\d+)\](.+)', header)
                
                if radio_match:
                    group_num = radio_match.group(1)
                    full_description = radio_match.group(2).strip()
                    
                    # Extract the common description (before the colon)
                    parts = full_description.split(':', 1)
                    if len(parts) == 2:
                        base_description = parts[0].strip()
                        option_text = parts[1].strip()
                    else:
                        base_description = full_description
                        option_text = full_description
                    
                    # Get point value
                    try:
                        points = float(point_values_row[col_idx]) if col_idx < len(point_values_row) else 0.0
                    except (ValueError, IndexError):
                        points = 0.0
                    
                    if group_num not in radio_groups:
                        radio_groups[group_num] = {
                            'description': base_description,
                            'options': {},
                            'max_points': 0.0,
                            'column_indices': [],
                            'column_to_letter': {}
                        }
                    
                    # Assign QWERTY letter to this option
                    current_option_count = len(radio_groups[group_num]['options'])
                    if current_option_count < len(QWERTY_LETTERS):
                        option_letter = QWERTY_LETTERS[current_option_count]
                        radio_groups[group_num]['options'][option_letter] = {
                            'text': option_text,
                            'points': str(points)
                        }
                        radio_groups[group_num]['max_points'] = max(radio_groups[group_num]['max_points'], points)
                        radio_groups[group_num]['column_indices'].append(col_idx)
                        # Store mapping from column index to letter for later use
                        radio_groups[group_num]['column_to_letter'] = radio_groups[group_num].get('column_to_letter', {})
                        radio_groups[group_num]['column_to_letter'][col_idx] = option_letter
                
                else:
                    # Regular checkbox item
                    try:
                        points = float(point_values_row[col_idx]) if col_idx < len(point_values_row) else 0.0
                    except (ValueError, IndexError):
                        points = 0.0
                        
                    rubric_item = RubricItem(
                        id=f"checkbox_{col_idx}",
                        description=header,
                        points=points,
                        type="CHECKBOX",
                        column_index=col_idx
                    )
                    rubric_items.append(rubric_item)
            
            # Add radio groups as rubric items
            for group_num, group_data in radio_groups.items():
                rubric_item = RubricItem(
                    id=f"radio_{group_num}",
                    description=group_data['description'],
                    points=group_data['max_points'],
                    type="RADIO",
                    options=group_data['options'],
                    column_index=min(group_data['column_indices'])  # Use first column index
                )
                rubric_items.append(rubric_item)
            
            # Parse student grades (skip header and last two rows)
            for row_idx in range(1, len(rows) - 2):
                row = rows[row_idx]
                if len(row) < 8:
                    continue
                    
                submission_id = row[0]
                name = row[2] if len(row) > 2 else ""
                
                if not submission_id or submission_id == "":
                    continue
                
                grades = {}
                
                # Process checkbox items
                for rubric_item in rubric_items:
                    if rubric_item.type == "CHECKBOX":
                        col_idx = rubric_item.column_index
                        if col_idx < len(row):
                            value = row[col_idx].strip().upper()
                            grades[rubric_item.id] = value == "TRUE"
                
                # Process radio items
                for group_num, group_data in radio_groups.items():
                    radio_id = f"radio_{group_num}"
                    selected_letter = None
                    
                    for col_idx in group_data['column_indices']:
                        if col_idx < len(row) and row[col_idx].strip().upper() == "TRUE":
                            # Get the letter for this column
                            selected_letter = group_data['column_to_letter'].get(col_idx)
                            break
                    
                    if selected_letter:
                        grades[radio_id] = selected_letter
                
                student_grades[submission_id] = StudentGrade(
                    submission_id=submission_id,
                    name=name,
                    grades=grades
                )
        
        return rubric_items, student_grades


class BackendClient:
    """Client for interacting with the grading backend."""
    
    def __init__(self, base_url: str = "http://localhost:8000"):
        self.base_url = base_url
        self.session = None
        
    async def __aenter__(self):
        self.session = aiohttp.ClientSession()
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            await self.session.close()
            
    async def grade_submission(
        self,
        assignment_context: Dict[str, str],
        source_files: Dict[str, str],
        rubric_items: List[Dict[str, Any]]
    ) -> List[EvaluationResult]:
        """Send grading request to backend and parse results."""
        
        request_data = {
            "assignment_context": assignment_context,
            "source_files": source_files,
            "rubric_items": rubric_items
        }
        
        # Log the complete request for debugging
        logger.info(f"=== SENDING REQUEST TO BACKEND ===")
        logger.info(f"URL: {self.base_url}/api/v1/grade-submission")
        logger.info(f"Assignment Context: {json.dumps(assignment_context, indent=2)}")
        logger.info(f"Number of source files: {len(source_files)}")
        logger.info(f"Source files: {list(source_files.keys())}")
        logger.info(f"Number of rubric items: {len(rubric_items)}")
        logger.info(f"Rubric items structure:")
        for i, item in enumerate(rubric_items):
            logger.info(f"  Item {i+1}: {json.dumps(item, indent=4)}")
        
        # Log complete request JSON (truncated source files for readability)
        debug_request = request_data.copy()
        debug_request["source_files"] = {
            filename: f"[{len(content)} characters]" if len(content) > 200 
            else content for filename, content in source_files.items()
        }
        logger.info(f"Complete request JSON structure:")
        logger.info(json.dumps(debug_request, indent=2))
        
        results = []
        
        try:
            async with self.session.post(
                f"{self.base_url}/api/v1/grade-submission",
                json=request_data,
                timeout=aiohttp.ClientTimeout(total=300)  # 5 minute timeout
            ) as response:
                logger.info(f"=== BACKEND RESPONSE ===")
                logger.info(f"Status: {response.status}")
                logger.info(f"Headers: {dict(response.headers)}")
                
                if response.status != 200:
                    error_text = await response.text()
                    logger.error(f"Backend error response: {error_text}")
                    raise Exception(f"Backend returned status {response.status}: {error_text}")
                
                logger.info("Starting to parse Server-Sent Events...")
                # Parse Server-Sent Events
                async for line in response.content:
                    line = line.decode('utf-8').strip()
                    logger.debug(f"Received SSE line: {line}")
                    if line.startswith('data: '):
                        data_str = line[6:]  # Remove 'data: ' prefix
                        try:
                            data = json.loads(data_str)
                            logger.info(f"Parsed SSE data: {json.dumps(data, indent=2)}")
                            if data.get('type') == 'partial_result' and data.get('decision'):
                                decision = data['decision']
                                verdict = decision.get('verdict', {})
                                
                                # Extract decision based on type
                                if decision['type'] == 'CHECKBOX':
                                    result_decision = verdict.get('decision', 'uncheck')
                                else:  # RADIO
                                    result_decision = verdict.get('selected_option', '')
                                
                                result = EvaluationResult(
                                    submission_id=assignment_context['submission_id'],
                                    rubric_id=decision['rubric_item_id'],
                                    decision=result_decision,
                                    confidence=verdict.get('confidence', 0) / 100.0,  # Convert to 0-1
                                    comment=verdict.get('comment', ''),
                                    evidence=verdict.get('evidence', {})
                                )
                                results.append(result)
                                logger.info(f"Added result for rubric {decision['rubric_item_id']}: {result_decision}")
                        except json.JSONDecodeError as e:
                            logger.warning(f"Failed to parse JSON from SSE data: {data_str}, error: {e}")
                            continue
                            
        except Exception as e:
            print(f"Error grading submission {assignment_context['submission_id']}: {str(e)}")
            
        return results


class ExcelReportGenerator:
    """Generates Excel reports comparing backend and human grades."""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        
    def create_report(
        self,
        project_name: str,
        rubric_items: List[RubricItem],
        human_grades: Dict[str, StudentGrade],
        backend_results: Dict[str, List[EvaluationResult]],
        output_path: Path
    ):
        """Create Excel report with comparison visualization."""
        
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = project_name[:31]  # Excel sheet name limit
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        false_positive_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        false_negative_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create headers
        headers = ["Submission ID", "Student Name"]
        
        # Add rubric item headers
        for rubric_item in rubric_items:
            headers.append(rubric_item.description)  # Full description without truncation
            headers.append("Backend Decision")
            headers.append("Confidence %")
            headers.append("Comment")
            headers.append("Match?")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Process each student
        row_idx = 2
        for submission_id, student_grade in human_grades.items():
            # Basic info
            self.ws.cell(row=row_idx, column=1, value=submission_id).border = border
            self.ws.cell(row=row_idx, column=2, value=student_grade.name).border = border
            
            # Get backend results for this student
            student_backend_results = backend_results.get(submission_id, [])
            backend_by_rubric = {r.rubric_id: r for r in student_backend_results}
            
            col_idx = 3
            for rubric_item in rubric_items:
                # Human grade
                human_value = student_grade.grades.get(rubric_item.id, False)
                
                if rubric_item.type == "CHECKBOX":
                    human_cell = self.ws.cell(row=row_idx, column=col_idx, value="TRUE" if human_value else "FALSE")
                else:  # RADIO
                    human_cell = self.ws.cell(row=row_idx, column=col_idx, value=str(human_value))
                human_cell.border = border
                
                # Backend result
                backend_result = backend_by_rubric.get(rubric_item.id)
                
                if backend_result:
                    if rubric_item.type == "CHECKBOX":
                        backend_value = "TRUE" if backend_result.decision == "check" else "FALSE"
                    else:  # RADIO
                        backend_value = backend_result.decision
                    
                    backend_cell = self.ws.cell(row=row_idx, column=col_idx + 1, value=backend_value)
                    confidence_cell = self.ws.cell(row=row_idx, column=col_idx + 2, value=f"{backend_result.confidence * 100:.1f}")
                    comment_cell = self.ws.cell(row=row_idx, column=col_idx + 3, value=backend_result.comment)
                    
                    # Check if match
                    if rubric_item.type == "CHECKBOX":
                        human_bool = human_value
                        backend_bool = backend_result.decision == "check"
                        match = human_bool == backend_bool
                        
                        if match:
                            match_text = "MATCH"
                            fill = match_fill
                        elif human_bool and not backend_bool:
                            match_text = "FALSE NEGATIVE"
                            fill = false_negative_fill
                        else:
                            match_text = "FALSE POSITIVE"
                            fill = false_positive_fill
                    else:  # RADIO
                        # For radio buttons, compare letters
                        match = str(human_value) == backend_result.decision
                        match_text = "MATCH" if match else "MISMATCH"
                        fill = match_fill if match else false_positive_fill
                    
                    match_cell = self.ws.cell(row=row_idx, column=col_idx + 4, value=match_text)
                    
                    # Apply formatting
                    for cell in [human_cell, backend_cell, match_cell]:
                        cell.fill = fill
                        cell.border = border
                    
                    confidence_cell.border = border
                    comment_cell.border = border
                    comment_cell.alignment = Alignment(wrap_text=True)
                else:
                    # No backend result
                    self.ws.cell(row=row_idx, column=col_idx + 1, value="N/A").border = border
                    self.ws.cell(row=row_idx, column=col_idx + 2, value="N/A").border = border
                    self.ws.cell(row=row_idx, column=col_idx + 3, value="No result").border = border
                    self.ws.cell(row=row_idx, column=col_idx + 4, value="NO DATA").border = border
                
                col_idx += 5
            
            row_idx += 1
        
        # Add summary statistics
        self.add_summary_sheet(rubric_items, human_grades, backend_results)
        
        # Adjust column widths
        for col in range(1, self.ws.max_column + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save
        self.wb.save(output_path)
        
    def add_summary_sheet(
        self,
        rubric_items: List[RubricItem],
        human_grades: Dict[str, StudentGrade],
        backend_results: Dict[str, List[EvaluationResult]]
    ):
        """Add summary statistics sheet."""
        
        summary_ws = self.wb.create_sheet("Summary")
        
        # Calculate statistics
        stats = []
        for rubric_item in rubric_items:
            total = 0
            matches = 0
            false_positives = 0
            false_negatives = 0
            no_data = 0
            avg_confidence = 0
            confidence_count = 0
            
            for submission_id, student_grade in human_grades.items():
                total += 1
                human_value = student_grade.grades.get(rubric_item.id, False)
                
                backend_results_student = backend_results.get(submission_id, [])
                backend_result = next((r for r in backend_results_student if r.rubric_id == rubric_item.id), None)
                
                if backend_result:
                    avg_confidence += backend_result.confidence
                    confidence_count += 1
                    
                    if rubric_item.type == "CHECKBOX":
                        human_bool = human_value
                        backend_bool = backend_result.decision == "check"
                        
                        if human_bool == backend_bool:
                            matches += 1
                        elif human_bool and not backend_bool:
                            false_negatives += 1
                        else:
                            false_positives += 1
                    else:  # RADIO
                        # For radio buttons, compare letters
                        if str(human_value) == backend_result.decision:
                            matches += 1
                        else:
                            false_positives += 1  # Treat all mismatches as false positives for radio
                else:
                    no_data += 1
            
            if confidence_count > 0:
                avg_confidence /= confidence_count
            
            stats.append({
                'rubric': rubric_item.description[:100],
                'type': rubric_item.type,
                'total': total,
                'matches': matches,
                'false_positives': false_positives,
                'false_negatives': false_negatives,
                'no_data': no_data,
                'accuracy': matches / (total - no_data) if (total - no_data) > 0 else 0,
                'avg_confidence': avg_confidence
            })
        
        # Write summary
        headers = ["Rubric Item", "Type", "Total", "Matches", "False Positives", "False Negatives", 
                   "No Data", "Accuracy %", "Avg Confidence %"]
        
        for col, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, stat in enumerate(stats, 2):
            summary_ws.cell(row=row, column=1, value=stat['rubric'])
            summary_ws.cell(row=row, column=2, value=stat['type'])
            summary_ws.cell(row=row, column=3, value=stat['total'])
            summary_ws.cell(row=row, column=4, value=stat['matches'])
            summary_ws.cell(row=row, column=5, value=stat['false_positives'])
            summary_ws.cell(row=row, column=6, value=stat['false_negatives'])
            summary_ws.cell(row=row, column=7, value=stat['no_data'])
            summary_ws.cell(row=row, column=8, value=f"{stat['accuracy'] * 100:.1f}")
            summary_ws.cell(row=row, column=9, value=f"{stat['avg_confidence'] * 100:.1f}")
        
        # Adjust column widths
        for col in range(1, 10):
            summary_ws.column_dimensions[get_column_letter(col)].width = 20


class CrossProjectUnifiedReportGenerator:
    """Generates a unified Excel report with statistics from all evaluated assignments."""
    
    def __init__(self):
        self.wb = None
    
    def create_cross_project_report(
        self,
        all_project_data: Dict[str, List[Dict]],
        output_path: Path
    ):
        """Create unified Excel report with statistics from all projects."""
        
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        subheader_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        excellent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        good_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue
        fair_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        poor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create overview summary sheet
        print("      📊 Creating overview summary sheet...")
        self._create_overview_sheet(all_project_data, header_fill, header_font, border)
        print("      ✅ Overview summary sheet completed")
        
        # Create detailed statistics sheet  
        print("      📋 Creating detailed statistics sheet...")
        self._create_detailed_stats_sheet(all_project_data, header_fill, header_font, subheader_fill, border)
        print("      ✅ Detailed statistics sheet completed")
        
        # Create project comparison sheet
        print("      📈 Creating project comparison sheet...")
        self._create_project_comparison_sheet(all_project_data, header_fill, header_font, border)
        print("      ✅ Project comparison sheet completed")
        
        # Create rubric analysis sheet
        print("      🎯 Creating rubric analysis sheet...")
        self._create_rubric_analysis_sheet(all_project_data, header_fill, header_font, border)
        print("      ✅ Rubric analysis sheet completed")
        
        # Save the workbook
        print("      💾 Saving Excel file...")
        self.wb.save(output_path)
        print("      ✅ Excel file saved successfully")
        
    def _create_overview_sheet(self, all_project_data, header_fill, header_font, border):
        """Create overview summary sheet with key statistics."""
        
        ws = self.wb.create_sheet("OVERVIEW_SUMMARY")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Cross-Project Evaluation Summary")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = [
            "Project", "Assignment", "Students Evaluated", "Total Rubric Items", 
            "Accuracy Rate", "Avg Confidence", "Checkbox Items", "Radio Items"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 4
        total_students = 0
        total_items = 0
        total_accuracy_sum = 0
        total_confidence_sum = 0
        project_count = 0
        
        for project_idx, (project_name, evaluation_data_list) in enumerate(all_project_data.items(), 1):
            print(f"         📊 Processing project {project_idx}/{len(all_project_data)}: {project_name}")
            for eval_data in evaluation_data_list:
                csv_name = eval_data['csv_name']
                rubric_items = eval_data['rubric_items']
                human_grades = eval_data['human_grades']
                backend_results = eval_data['backend_results']
                
                # Calculate statistics
                students_evaluated = len(human_grades)
                total_rubric_items = len(rubric_items)
                checkbox_items = len([r for r in rubric_items if r.type == "CHECKBOX"])
                radio_items = len([r for r in rubric_items if r.type == "RADIO"])
                
                # Calculate accuracy and confidence
                matches = 0
                total_comparisons = 0
                confidence_sum = 0
                confidence_count = 0
                
                for student_id, student_grade in human_grades.items():
                    backend_results_student = backend_results.get(student_id, [])
                    
                    # Create a lookup dictionary for this student's backend results to optimize performance
                    backend_lookup = {r.rubric_id: r for r in backend_results_student}
                    
                    for rubric_item in rubric_items:
                        backend_result = backend_lookup.get(rubric_item.id)
                        
                        if backend_result:
                            total_comparisons += 1
                            confidence_sum += backend_result.confidence
                            confidence_count += 1
                            
                            human_value = student_grade.grades.get(rubric_item.id, False)
                            
                            if rubric_item.type == "CHECKBOX":
                                backend_bool = backend_result.decision == "check"
                                if human_value == backend_bool:
                                    matches += 1
                            else:  # RADIO
                                if str(human_value) == backend_result.decision:
                                    matches += 1
                
                accuracy_rate = (matches / total_comparisons * 100) if total_comparisons > 0 else 0
                avg_confidence = (confidence_sum / confidence_count * 100) if confidence_count > 0 else 0
                
                # Write data
                data = [
                    project_name, csv_name, students_evaluated, total_rubric_items,
                    f"{accuracy_rate:.1f}%", f"{avg_confidence:.1f}%", checkbox_items, radio_items
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    
                    # Color code accuracy
                    if col == 5:  # Accuracy column
                        if accuracy_rate >= 90:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif accuracy_rate >= 80:
                            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                        elif accuracy_rate >= 70:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Accumulate totals
                total_students += students_evaluated
                total_items += total_rubric_items
                total_accuracy_sum += accuracy_rate
                total_confidence_sum += avg_confidence
                project_count += 1
                
                row += 1
        
        # Add summary row
        row += 1
        summary_headers = ["TOTALS/AVERAGES", "", f"{total_students}", f"{total_items}", 
                          f"{total_accuracy_sum/project_count:.1f}%" if project_count > 0 else "0%",
                          f"{total_confidence_sum/project_count:.1f}%" if project_count > 0 else "0%", "", ""]
        
        for col, value in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
    
    def _create_detailed_stats_sheet(self, all_project_data, header_fill, header_font, subheader_fill, border):
        """Create detailed statistics sheet with per-rubric-item analysis."""
        
        ws = self.wb.create_sheet("DETAILED_STATISTICS")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Detailed Rubric Item Analysis")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')
        
        # Headers
        headers = [
            "Project", "Assignment", "Rubric Item", "Type", "Points", 
            "Matches", "False Positives", "False Negatives", "Accuracy", "Avg Confidence"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 4
        
        for project_name, evaluation_data_list in all_project_data.items():
            for eval_data in evaluation_data_list:
                csv_name = eval_data['csv_name']
                rubric_items = eval_data['rubric_items']
                human_grades = eval_data['human_grades']
                backend_results = eval_data['backend_results']
                
                for rubric_item in rubric_items:
                    matches = 0
                    false_positives = 0
                    false_negatives = 0
                    total = 0
                    confidence_sum = 0
                    confidence_count = 0
                    
                    for student_id, student_grade in human_grades.items():
                        backend_results_student = backend_results.get(student_id, [])
                        backend_result = next((r for r in backend_results_student if r.rubric_id == rubric_item.id), None)
                        
                        if backend_result:
                            total += 1
                            confidence_sum += backend_result.confidence
                            confidence_count += 1
                            
                            human_value = student_grade.grades.get(rubric_item.id, False)
                            
                            if rubric_item.type == "CHECKBOX":
                                human_bool = human_value
                                backend_bool = backend_result.decision == "check"
                                
                                if human_bool == backend_bool:
                                    matches += 1
                                elif not human_bool and backend_bool:
                                    false_positives += 1
                                else:
                                    false_negatives += 1
                            else:  # RADIO
                                if str(human_value) == backend_result.decision:
                                    matches += 1
                                else:
                                    false_positives += 1
                    
                    accuracy = (matches / total * 100) if total > 0 else 0
                    avg_confidence = (confidence_sum / confidence_count * 100) if confidence_count > 0 else 0
                    
                    # Write data
                    data = [
                        project_name, csv_name, rubric_item.description[:50] + "..." if len(rubric_item.description) > 50 else rubric_item.description,
                        rubric_item.type, rubric_item.points, matches, false_positives, false_negatives,
                        f"{accuracy:.1f}%", f"{avg_confidence:.1f}%"
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        
                        # Color code accuracy
                        if col == 9:  # Accuracy column
                            if accuracy >= 90:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif accuracy >= 80:
                                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                            elif accuracy >= 70:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 12
    
    def _create_project_comparison_sheet(self, all_project_data, header_fill, header_font, border):
        """Create project comparison sheet showing summary by project."""
        
        ws = self.wb.create_sheet("PROJECT_COMPARISON")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Project Comparison Summary")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:I1')
        
        # Headers
        headers = [
            "Project", "Total Assignments", "Total Students", "Total Rubric Items",
            "Overall Accuracy", "Avg Confidence", "Best Assignment", "Worst Assignment", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 4
        
        for project_name, evaluation_data_list in all_project_data.items():
            total_assignments = len(evaluation_data_list)
            total_students = 0
            total_items = 0
            accuracy_scores = []
            confidence_scores = []
            assignment_accuracies = []
            
            for eval_data in evaluation_data_list:
                csv_name = eval_data['csv_name']
                rubric_items = eval_data['rubric_items']
                human_grades = eval_data['human_grades']
                backend_results = eval_data['backend_results']
                
                students_count = len(human_grades)
                items_count = len(rubric_items)
                total_students += students_count
                total_items += items_count
                
                # Calculate accuracy for this assignment
                matches = 0
                total_comparisons = 0
                confidence_sum = 0
                confidence_count = 0
                
                for student_id, student_grade in human_grades.items():
                    backend_results_student = backend_results.get(student_id, [])
                    
                    for rubric_item in rubric_items:
                        backend_result = next((r for r in backend_results_student if r.rubric_id == rubric_item.id), None)
                        
                        if backend_result:
                            total_comparisons += 1
                            confidence_sum += backend_result.confidence
                            confidence_count += 1
                            
                            human_value = student_grade.grades.get(rubric_item.id, False)
                            
                            if rubric_item.type == "CHECKBOX":
                                backend_bool = backend_result.decision == "check"
                                if human_value == backend_bool:
                                    matches += 1
                            else:  # RADIO
                                if str(human_value) == backend_result.decision:
                                    matches += 1
                
                assignment_accuracy = (matches / total_comparisons * 100) if total_comparisons > 0 else 0
                assignment_confidence = (confidence_sum / confidence_count * 100) if confidence_count > 0 else 0
                
                accuracy_scores.append(assignment_accuracy)
                confidence_scores.append(assignment_confidence)
                assignment_accuracies.append((csv_name, assignment_accuracy))
            
            # Calculate overall statistics
            overall_accuracy = sum(accuracy_scores) / len(accuracy_scores) if accuracy_scores else 0
            avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0
            
            # Find best and worst assignments
            assignment_accuracies.sort(key=lambda x: x[1], reverse=True)
            best_assignment = assignment_accuracies[0][0] if assignment_accuracies else "N/A"
            worst_assignment = assignment_accuracies[-1][0] if assignment_accuracies else "N/A"
            
            # Generate notes
            notes = []
            if overall_accuracy >= 90:
                notes.append("Excellent performance")
            elif overall_accuracy >= 80:
                notes.append("Good performance")
            elif overall_accuracy >= 70:
                notes.append("Fair performance")
            else:
                notes.append("Needs improvement")
            
            if avg_confidence >= 85:
                notes.append("High confidence")
            elif avg_confidence >= 70:
                notes.append("Moderate confidence")
            else:
                notes.append("Low confidence")
            
            # Write data
            data = [
                project_name, total_assignments, total_students, total_items,
                f"{overall_accuracy:.1f}%", f"{avg_confidence:.1f}%", 
                best_assignment, worst_assignment, "; ".join(notes)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Color code overall accuracy
                if col == 5:  # Overall accuracy column
                    if overall_accuracy >= 90:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif overall_accuracy >= 80:
                        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    elif overall_accuracy >= 70:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 30
    
    def _create_rubric_analysis_sheet(self, all_project_data, header_fill, header_font, border):
        """Create rubric type analysis sheet."""
        
        ws = self.wb.create_sheet("RUBRIC_TYPE_ANALYSIS")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="Rubric Type Performance Analysis")
        title_cell.font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Collect statistics by rubric type
        checkbox_stats = []
        radio_stats = []
        
        for project_name, evaluation_data_list in all_project_data.items():
            for eval_data in evaluation_data_list:
                csv_name = eval_data['csv_name']
                rubric_items = eval_data['rubric_items']
                human_grades = eval_data['human_grades']
                backend_results = eval_data['backend_results']
                
                checkbox_matches = 0
                checkbox_total = 0
                checkbox_confidence = 0
                checkbox_count = 0
                
                radio_matches = 0
                radio_total = 0
                radio_confidence = 0
                radio_count = 0
                
                for rubric_item in rubric_items:
                    for student_id, student_grade in human_grades.items():
                        backend_results_student = backend_results.get(student_id, [])
                        backend_result = next((r for r in backend_results_student if r.rubric_id == rubric_item.id), None)
                        
                        if backend_result:
                            human_value = student_grade.grades.get(rubric_item.id, False)
                            
                            if rubric_item.type == "CHECKBOX":
                                checkbox_total += 1
                                checkbox_confidence += backend_result.confidence
                                checkbox_count += 1
                                
                                backend_bool = backend_result.decision == "check"
                                if human_value == backend_bool:
                                    checkbox_matches += 1
                            
                            else:  # RADIO
                                radio_total += 1
                                radio_confidence += backend_result.confidence
                                radio_count += 1
                                
                                if str(human_value) == backend_result.decision:
                                    radio_matches += 1
                
                if checkbox_total > 0:
                    checkbox_stats.append({
                        'project': project_name,
                        'assignment': csv_name,
                        'accuracy': checkbox_matches / checkbox_total * 100,
                        'confidence': checkbox_confidence / checkbox_count * 100,
                        'total': checkbox_total
                    })
                
                if radio_total > 0:
                    radio_stats.append({
                        'project': project_name,
                        'assignment': csv_name,
                        'accuracy': radio_matches / radio_total * 100,
                        'confidence': radio_confidence / radio_count * 100,
                        'total': radio_total
                    })
        
        # Headers for checkbox analysis
        row = 3
        ws.cell(row=row, column=1, value="CHECKBOX ITEMS ANALYSIS").font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Project", "Assignment", "Accuracy", "Avg Confidence", "Total Items", "Performance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row += 1
        for stat in checkbox_stats:
            performance = "Excellent" if stat['accuracy'] >= 90 else "Good" if stat['accuracy'] >= 80 else "Fair" if stat['accuracy'] >= 70 else "Poor"
            data = [stat['project'], stat['assignment'], f"{stat['accuracy']:.1f}%", 
                   f"{stat['confidence']:.1f}%", stat['total'], performance]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                if col == 3:  # Accuracy column
                    if stat['accuracy'] >= 90:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif stat['accuracy'] >= 80:
                        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    elif stat['accuracy'] >= 70:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Add spacing and radio analysis
        row += 2
        ws.cell(row=row, column=1, value="RADIO ITEMS ANALYSIS").font = Font(bold=True, size=12)
        row += 1
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row += 1
        for stat in radio_stats:
            performance = "Excellent" if stat['accuracy'] >= 90 else "Good" if stat['accuracy'] >= 80 else "Fair" if stat['accuracy'] >= 70 else "Poor"
            data = [stat['project'], stat['assignment'], f"{stat['accuracy']:.1f}%", 
                   f"{stat['confidence']:.1f}%", stat['total'], performance]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                if col == 3:  # Accuracy column
                    if stat['accuracy'] >= 90:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif stat['accuracy'] >= 80:
                        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    elif stat['accuracy'] >= 70:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15


class UnifiedExcelReportGenerator:
    """Generates unified Excel reports combining multiple CSV evaluations."""
    
    def __init__(self):
        self.wb = None
    
    def create_unified_report(
        self,
        project_name: str,
        all_evaluation_data: List[Dict],
        output_path: Path
    ):
        """Create unified Excel report with separate sheets for each CSV and combined summary."""
        
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        false_positive_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        false_negative_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create sheet for each CSV evaluation
        for eval_data in all_evaluation_data:
            csv_name = eval_data['csv_name']
            rubric_items = eval_data['rubric_items']
            human_grades = eval_data['human_grades']
            backend_results = eval_data['backend_results']
            
            # Create sheet (truncate name if too long for Excel)
            sheet_name = csv_name[:31] if len(csv_name) > 31 else csv_name
            ws = self.wb.create_sheet(title=sheet_name)
            
            # Create headers
            headers = ["Submission ID", "Student Name"]
            for rubric_item in rubric_items:
                headers.append(rubric_item.description)  # Full description
                headers.append("Backend Decision")
                headers.append("Confidence %")
                headers.append("Comment")
                headers.append("Match?")
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Process each student
            row_idx = 2
            for submission_id, student_grade in human_grades.items():
                # Basic info
                ws.cell(row=row_idx, column=1, value=submission_id).border = border
                ws.cell(row=row_idx, column=2, value=student_grade.name).border = border
                
                # Get backend results for this student
                student_backend_results = backend_results.get(submission_id, [])
                backend_by_rubric = {r.rubric_id: r for r in student_backend_results}
                
                col_idx = 3
                for rubric_item in rubric_items:
                    # Human grade
                    human_value = student_grade.grades.get(rubric_item.id, False)
                    
                    if rubric_item.type == "CHECKBOX":
                        human_cell = ws.cell(row=row_idx, column=col_idx, value="TRUE" if human_value else "FALSE")
                    else:  # RADIO
                        human_cell = ws.cell(row=row_idx, column=col_idx, value=str(human_value))
                    human_cell.border = border
                    
                    # Backend result
                    backend_result = backend_by_rubric.get(rubric_item.id)
                    
                    if backend_result:
                        if rubric_item.type == "CHECKBOX":
                            backend_value = "TRUE" if backend_result.decision == "check" else "FALSE"
                        else:  # RADIO
                            backend_value = backend_result.decision
                        
                        backend_cell = ws.cell(row=row_idx, column=col_idx + 1, value=backend_value)
                        confidence_cell = ws.cell(row=row_idx, column=col_idx + 2, value=f"{backend_result.confidence * 100:.1f}")
                        comment_cell = ws.cell(row=row_idx, column=col_idx + 3, value=backend_result.comment)
                        
                        # Check if match
                        if rubric_item.type == "CHECKBOX":
                            human_bool = human_value
                            backend_bool = backend_result.decision == "check"
                            match = human_bool == backend_bool
                            
                            if match:
                                match_text = "MATCH"
                                fill = match_fill
                            elif human_bool and not backend_bool:
                                match_text = "FALSE NEGATIVE"
                                fill = false_negative_fill
                            else:
                                match_text = "FALSE POSITIVE"
                                fill = false_positive_fill
                        else:  # RADIO
                            # For radio buttons, compare letters
                            match = str(human_value) == backend_result.decision
                            match_text = "MATCH" if match else "MISMATCH"
                            fill = match_fill if match else false_positive_fill
                        
                        match_cell = ws.cell(row=row_idx, column=col_idx + 4, value=match_text)
                        
                        # Apply formatting
                        for cell in [human_cell, backend_cell, match_cell]:
                            cell.fill = fill
                            cell.border = border
                        
                        confidence_cell.border = border
                        comment_cell.border = border
                        comment_cell.alignment = Alignment(wrap_text=True)
                    else:
                        # No backend result
                        ws.cell(row=row_idx, column=col_idx + 1, value="N/A").border = border
                        ws.cell(row=row_idx, column=col_idx + 2, value="N/A").border = border
                        ws.cell(row=row_idx, column=col_idx + 3, value="No result").border = border
                        ws.cell(row=row_idx, column=col_idx + 4, value="NO DATA").border = border
                    
                    col_idx += 5
                
                row_idx += 1
            
            # Adjust column widths
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Create comprehensive summary sheet
        self.create_comprehensive_summary_sheet(all_evaluation_data)
        
        # Create points comparison sheets
        self.create_points_comparison_sheets(all_evaluation_data)
        
        # Save the workbook
        self.wb.save(output_path)
    
    def create_points_comparison_sheets(self, all_evaluation_data: List[Dict]):
        """Create sheets with points comparison analysis."""
        
        # Collect all points comparisons and stats across all CSVs
        all_comparisons = []
        all_rubric_stats = []
        
        # Convert backend results to the format expected by PointsCalculator
        for eval_data in all_evaluation_data:
            csv_name = eval_data['csv_name']
            rubric_items = eval_data['rubric_items']
            human_grades = eval_data['human_grades']
            backend_results = eval_data['backend_results']
            
            # Convert backend results to AIGradingResult format
            ai_results = {}
            for submission_id, results_list in backend_results.items():
                ai_decisions = {}
                for result in results_list:
                    # Create mock decision object
                    decision = type('obj', (object,), {
                        'verdict': type('obj', (object,), {
                            'decision': type('obj', (object,), {'value': 'CHECK' if result.decision == 'check' else 'CROSS'})() if result.rubric_id.startswith('checkbox') else None,
                            'selected_option': result.decision if result.rubric_id.startswith('radio') else None
                        })()
                    })()
                    ai_decisions[result.rubric_id] = decision
                
                ai_results[submission_id] = AIGradingResult(
                    submission_id=submission_id,
                    decisions=ai_decisions
                )
            
            # Calculate points comparisons for this CSV
            if ai_results:
                comparisons = PointsCalculator.compare_submissions(human_grades, ai_results, rubric_items)
                stats = PointsCalculator.calculate_rubric_stats(comparisons, rubric_items)
                
                # Add CSV identifier to each comparison and stat
                for comp in comparisons:
                    comp.csv_name = csv_name
                for stat in stats:
                    stat.csv_name = csv_name
                
                all_comparisons.extend(comparisons)
                all_rubric_stats.extend(stats)
        
        if not all_comparisons:
            return
        
        # Create submissions points comparison sheet
        self.create_submissions_points_sheet(all_comparisons)
        
        # Create rubric items points analysis sheet
        self.create_rubric_points_analysis_sheet(all_rubric_stats)
        
        # Create overall points summary sheet
        self.create_points_summary_sheet(all_comparisons, all_rubric_stats)
    
    def create_submissions_points_sheet(self, comparisons: List[PointsComparison]):
        """Create sheet showing points comparison for each submission."""
        
        ws = self.wb.create_sheet("POINTS_BY_SUBMISSION")
        
        # Define styles
        header_fill = PatternFill(start_color="0F243E", end_color="0F243E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for AI higher
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for AI lower
        neutral_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Blue for same
        
        # Headers
        headers = ["CSV File", "Submission ID", "Human Total", "AI Total", "Difference (AI-Human)", "% Difference", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Sort comparisons by absolute difference (largest discrepancies first)
        sorted_comparisons = sorted(comparisons, key=lambda x: abs(x.difference), reverse=True)
        
        # Fill data
        for row, comp in enumerate(sorted_comparisons, 2):
            # Calculate percentage difference
            if comp.human_total > 0:
                pct_diff = (comp.difference / comp.human_total) * 100
            else:
                pct_diff = 0.0 if comp.ai_total == 0 else float('inf')
            
            # Determine status and color
            if abs(comp.difference) < 0.01:
                status = "EXACT MATCH"
                fill = neutral_fill
            elif comp.difference > 0:
                status = "AI HIGHER"
                fill = positive_fill
            else:
                status = "AI LOWER"
                fill = negative_fill
            
            # Fill row data
            ws.cell(row=row, column=1, value=getattr(comp, 'csv_name', 'Unknown'))
            ws.cell(row=row, column=2, value=comp.submission_id)
            ws.cell(row=row, column=3, value=comp.human_total).number_format = '0.0'
            ws.cell(row=row, column=4, value=comp.ai_total).number_format = '0.0'
            ws.cell(row=row, column=5, value=comp.difference).number_format = '0.0'
            
            pct_cell = ws.cell(row=row, column=6, value=pct_diff if pct_diff != float('inf') else 'N/A')
            if pct_diff != float('inf'):
                pct_cell.number_format = '0.0"%"'
            
            status_cell = ws.cell(row=row, column=7, value=status)
            
            # Apply formatting to difference and status cells
            for col in [5, 6, 7]:
                ws.cell(row=row, column=col).fill = fill
        
        # Adjust column widths
        column_widths = [15, 15, 12, 12, 15, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def create_rubric_points_analysis_sheet(self, stats: List[RubricItemStats]):
        """Create sheet showing points analysis for each rubric item."""
        
        ws = self.wb.create_sheet("RUBRIC_POINTS_ANALYSIS")
        
        # Define styles
        header_fill = PatternFill(start_color="0F243E", end_color="0F243E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = [
            "CSV File", "Rubric Item", "Type", "Max Points", 
            "Human Avg", "AI Avg", "Difference", "Human Std", "AI Std",
            "Agreement %", "Sample Size"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Sort stats by absolute difference (most problematic first)
        sorted_stats = sorted(stats, key=lambda x: abs(x.ai_avg - x.human_avg), reverse=True)
        
        # Fill data
        for row, stat in enumerate(sorted_stats, 2):
            difference = stat.ai_avg - stat.human_avg
            
            ws.cell(row=row, column=1, value=getattr(stat, 'csv_name', 'Unknown'))
            ws.cell(row=row, column=2, value=stat.description[:50] + "..." if len(stat.description) > 50 else stat.description)
            ws.cell(row=row, column=3, value=stat.rubric_type)
            ws.cell(row=row, column=4, value=stat.max_points).number_format = '0.0'
            ws.cell(row=row, column=5, value=stat.human_avg).number_format = '0.00'
            ws.cell(row=row, column=6, value=stat.ai_avg).number_format = '0.00'
            ws.cell(row=row, column=7, value=difference).number_format = '0.00'
            ws.cell(row=row, column=8, value=stat.human_std).number_format = '0.00'
            ws.cell(row=row, column=9, value=stat.ai_std).number_format = '0.00'
            ws.cell(row=row, column=10, value=stat.agreement_rate).number_format = '0.0"%"'
            ws.cell(row=row, column=11, value=stat.sample_size)
            
            # Color code the difference column
            diff_cell = ws.cell(row=row, column=7)
            if abs(difference) < 0.01:
                diff_cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Blue
            elif difference > 0:
                diff_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            else:
                diff_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        
        # Adjust column widths
        column_widths = [15, 40, 10, 10, 10, 10, 12, 10, 10, 12, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def create_points_summary_sheet(self, comparisons: List[PointsComparison], stats: List[RubricItemStats]):
        """Create sheet with overall points analysis summary."""
        
        ws = self.wb.create_sheet("POINTS_SUMMARY", 0)  # Make it the first sheet
        
        # Calculate overall statistics
        if not comparisons:
            ws.cell(row=1, column=1, value="No points comparison data available")
            return
            
        human_totals = [c.human_total for c in comparisons]
        ai_totals = [c.ai_total for c in comparisons]
        differences = [c.difference for c in comparisons]
        
        human_avg = statistics.mean(human_totals)
        ai_avg = statistics.mean(ai_totals)
        human_std = statistics.stdev(human_totals) if len(human_totals) > 1 else 0.0
        ai_std = statistics.stdev(ai_totals) if len(ai_totals) > 1 else 0.0
        
        avg_diff = statistics.mean(differences)
        abs_avg_diff = statistics.mean([abs(d) for d in differences])
        
        # Agreement statistics
        exact_matches = sum(1 for c in comparisons if abs(c.difference) < 0.01)
        close_matches = sum(1 for c in comparisons if abs(c.difference) <= 1.0)
        
        # Write summary
        row = 1
        
        # Title
        title_cell = ws.cell(row=row, column=1, value="📊 POINTS ANALYSIS SUMMARY")
        title_cell.font = Font(bold=True, size=16, color="0F243E")
        row += 2
        
        # Basic statistics
        ws.cell(row=row, column=1, value="OVERALL STATISTICS").font = Font(bold=True, size=12)
        row += 1
        
        summary_data = [
            ("Sample Size:", f"{len(comparisons)} submissions"),
            ("", ""),
            ("Human Average:", f"{human_avg:.2f} ± {human_std:.2f} points"),
            ("AI Average:", f"{ai_avg:.2f} ± {ai_std:.2f} points"),
            ("Average Difference:", f"{avg_diff:.2f} points (AI - Human)"),
            ("Avg Absolute Difference:", f"{abs_avg_diff:.2f} points"),
            ("Percentage Difference:", f"{(avg_diff / human_avg * 100):.1f}%" if human_avg > 0 else "N/A"),
            ("", ""),
            ("Exact Matches:", f"{exact_matches}/{len(comparisons)} ({exact_matches/len(comparisons)*100:.1f}%)"),
            ("Close Matches (±1pt):", f"{close_matches}/{len(comparisons)} ({close_matches/len(comparisons)*100:.1f}%)"),
        ]
        
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True if label and not label.startswith(" ") else False)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Rubric-level summary
        if stats:
            row += 1
            ws.cell(row=row, column=1, value="RUBRIC ITEM ANALYSIS").font = Font(bold=True, size=12)
            row += 1
            
            avg_agreement = statistics.mean([s.agreement_rate for s in stats])
            ws.cell(row=row, column=1, value="Average Agreement Rate:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{avg_agreement:.1f}%")
            row += 2
            
            # Most problematic rubric items
            problematic = sorted(stats, key=lambda s: abs(s.ai_avg - s.human_avg), reverse=True)[:5]
            ws.cell(row=row, column=1, value="MOST DIFFERENT RUBRIC ITEMS").font = Font(bold=True, size=11)
            row += 1
            
            for i, stat in enumerate(problematic, 1):
                diff = stat.ai_avg - stat.human_avg
                desc = stat.description[:60] + "..." if len(stat.description) > 60 else stat.description
                
                ws.cell(row=row, column=1, value=f"{i}. {desc}").font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value=f"   Human: {stat.human_avg:.2f}, AI: {stat.ai_avg:.2f}, Diff: {diff:.2f}")
                
                # Color code the difference
                diff_cell = ws.cell(row=row, column=1)
                if abs(diff) > 1.0:
                    diff_cell.font = Font(color="C00000")  # Red for large differences
                elif abs(diff) > 0.5:
                    diff_cell.font = Font(color="FF8C00")  # Orange for medium differences
                
                row += 1
        
        # Distribution analysis
        row += 1
        ws.cell(row=row, column=1, value="DISTRIBUTION ANALYSIS").font = Font(bold=True, size=12)
        row += 1
        
        # Count submissions by difference ranges
        large_pos = sum(1 for c in comparisons if c.difference >= 2.0)
        small_pos = sum(1 for c in comparisons if 0.1 <= c.difference < 2.0)
        exact = sum(1 for c in comparisons if abs(c.difference) < 0.1)
        small_neg = sum(1 for c in comparisons if -2.0 < c.difference <= -0.1)
        large_neg = sum(1 for c in comparisons if c.difference <= -2.0)
        
        distribution_data = [
            ("AI Much Higher (≥2pts):", f"{large_pos} ({large_pos/len(comparisons)*100:.1f}%)"),
            ("AI Slightly Higher:", f"{small_pos} ({small_pos/len(comparisons)*100:.1f}%)"),
            ("Same Score (±0.1pts):", f"{exact} ({exact/len(comparisons)*100:.1f}%)"),
            ("AI Slightly Lower:", f"{small_neg} ({small_neg/len(comparisons)*100:.1f}%)"),
            ("AI Much Lower (≤-2pts):", f"{large_neg} ({large_neg/len(comparisons)*100:.1f}%)"),
        ]
        
        for label, value in distribution_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
    
    def create_comprehensive_summary_sheet(self, all_evaluation_data: List[Dict]):
        """Create a comprehensive summary sheet combining all CSV evaluations."""
        
        summary_ws = self.wb.create_sheet("COMPREHENSIVE_SUMMARY")
        
        # Calculate combined statistics
        all_stats = []
        
        for eval_data in all_evaluation_data:
            csv_name = eval_data['csv_name']
            rubric_items = eval_data['rubric_items']
            human_grades = eval_data['human_grades']
            backend_results = eval_data['backend_results']
            
            for rubric_item in rubric_items:
                total = 0
                matches = 0
                false_positives = 0
                false_negatives = 0
                no_data = 0
                avg_confidence = 0
                confidence_count = 0
                
                for submission_id, student_grade in human_grades.items():
                    total += 1
                    human_value = student_grade.grades.get(rubric_item.id, False)
                    
                    backend_results_student = backend_results.get(submission_id, [])
                    backend_result = next((r for r in backend_results_student if r.rubric_id == rubric_item.id), None)
                    
                    if backend_result:
                        avg_confidence += backend_result.confidence
                        confidence_count += 1
                        
                        if rubric_item.type == "CHECKBOX":
                            human_bool = human_value
                            backend_bool = backend_result.decision == "check"
                            
                            if human_bool == backend_bool:
                                matches += 1
                            elif human_bool and not backend_bool:
                                false_negatives += 1
                            else:
                                false_positives += 1
                        else:  # RADIO
                            # For radio buttons, compare letters
                            if str(human_value) == backend_result.decision:
                                matches += 1
                            else:
                                false_positives += 1  # Treat all mismatches as false positives for radio
                    else:
                        no_data += 1
                
                if confidence_count > 0:
                    avg_confidence /= confidence_count
                
                all_stats.append({
                    'csv_name': csv_name,
                    'rubric': rubric_item.description,
                    'type': rubric_item.type,
                    'total': total,
                    'matches': matches,
                    'false_positives': false_positives,
                    'false_negatives': false_negatives,
                    'no_data': no_data,
                    'accuracy': matches / (total - no_data) if (total - no_data) > 0 else 0,
                    'avg_confidence': avg_confidence
                })
        
        # Write comprehensive summary
        headers = ["CSV File", "Rubric Item", "Type", "Total", "Matches", "False Positives", 
                   "False Negatives", "No Data", "Accuracy %", "Avg Confidence %"]
        
        for col, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, stat in enumerate(all_stats, 2):
            summary_ws.cell(row=row, column=1, value=stat['csv_name'])
            summary_ws.cell(row=row, column=2, value=stat['rubric'])
            summary_ws.cell(row=row, column=3, value=stat['type'])
            summary_ws.cell(row=row, column=4, value=stat['total'])
            summary_ws.cell(row=row, column=5, value=stat['matches'])
            summary_ws.cell(row=row, column=6, value=stat['false_positives'])
            summary_ws.cell(row=row, column=7, value=stat['false_negatives'])
            summary_ws.cell(row=row, column=8, value=stat['no_data'])
            summary_ws.cell(row=row, column=9, value=f"{stat['accuracy'] * 100:.1f}")
            summary_ws.cell(row=row, column=10, value=f"{stat['avg_confidence'] * 100:.1f}")
        
        # Adjust column widths for summary
        for col in range(1, 11):
            summary_ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Add overall project statistics at the top after some empty rows
        summary_ws.insert_rows(1, 5)
        
        # Calculate overall project stats
        total_evaluations = sum(stat['total'] for stat in all_stats)
        total_matches = sum(stat['matches'] for stat in all_stats)
        total_false_positives = sum(stat['false_positives'] for stat in all_stats)
        total_false_negatives = sum(stat['false_negatives'] for stat in all_stats)
        total_no_data = sum(stat['no_data'] for stat in all_stats)
        
        overall_accuracy = total_matches / (total_evaluations - total_no_data) if (total_evaluations - total_no_data) > 0 else 0
        overall_confidence = sum(stat['avg_confidence'] * stat['total'] for stat in all_stats) / total_evaluations if total_evaluations > 0 else 0
        
        # Write project overview
        summary_ws.cell(row=1, column=1, value="PROJECT OVERVIEW").font = Font(bold=True, size=14)
        summary_ws.cell(row=2, column=1, value=f"Total Evaluations: {total_evaluations}")
        summary_ws.cell(row=2, column=3, value=f"Overall Accuracy: {overall_accuracy * 100:.1f}%")
        summary_ws.cell(row=3, column=1, value=f"Total Matches: {total_matches}")
        summary_ws.cell(row=3, column=3, value=f"Overall Confidence: {overall_confidence * 100:.1f}%")
        summary_ws.cell(row=4, column=1, value=f"False Positives: {total_false_positives}")
        summary_ws.cell(row=4, column=3, value=f"False Negatives: {total_false_negatives}")


class PointsReporter:
    """Generates reports comparing AI and human grading points."""
    
    @staticmethod
    def print_submission_comparison(comparisons: List[PointsComparison], limit: int = 10) -> None:
        """Print detailed submission-by-submission comparison."""
        print(f"\n{'='*80}")
        print("📊 SUBMISSION-BY-SUBMISSION POINTS COMPARISON")
        print(f"{'='*80}")
        print(f"{'Submission ID':<15} {'Human Total':<12} {'AI Total':<12} {'Difference':<12} {'% Diff':<12}")
        print("-" * 80)
        
        # Sort by absolute difference (largest discrepancies first)
        sorted_comparisons = sorted(comparisons, key=lambda x: abs(x.difference), reverse=True)
        
        for i, comp in enumerate(sorted_comparisons[:limit]):
            if comp.human_total > 0:
                pct_diff = (comp.difference / comp.human_total) * 100
            else:
                pct_diff = 0.0 if comp.ai_total == 0 else float('inf')
            
            print(f"{comp.submission_id:<15} {comp.human_total:<12.1f} {comp.ai_total:<12.1f} "
                  f"{comp.difference:<12.1f} {pct_diff:<12.1f}%")
        
        if len(sorted_comparisons) > limit:
            print(f"... and {len(sorted_comparisons) - limit} more submissions")
    
    @staticmethod
    def print_rubric_stats(stats: List[RubricItemStats]) -> None:
        """Print detailed rubric item statistics."""
        print(f"\n{'='*100}")
        print("📋 RUBRIC ITEM ANALYSIS - HUMAN vs AI AVERAGES")
        print(f"{'='*100}")
        print(f"{'Rubric Item':<40} {'Type':<10} {'Max Pts':<8} {'Human Avg':<10} {'AI Avg':<10} "
              f"{'Diff':<8} {'Agreement':<10} {'Samples':<8}")
        print("-" * 100)
        
        for stat in stats:
            diff = stat.ai_avg - stat.human_avg
            description = stat.description[:37] + "..." if len(stat.description) > 40 else stat.description
            
            print(f"{description:<40} {stat.rubric_type:<10} {stat.max_points:<8.1f} "
                  f"{stat.human_avg:<10.2f} {stat.ai_avg:<10.2f} {diff:<8.2f} "
                  f"{stat.agreement_rate:<10.1f}% {stat.sample_size:<8}")
    
    @staticmethod
    def print_overall_summary(comparisons: List[PointsComparison], stats: List[RubricItemStats]) -> None:
        """Print overall summary statistics."""
        if not comparisons:
            print("No comparison data available")
            return
            
        print(f"\n{'='*60}")
        print("📈 OVERALL GRADING COMPARISON SUMMARY")
        print(f"{'='*60}")
        
        # Calculate overall statistics
        human_totals = [c.human_total for c in comparisons]
        ai_totals = [c.ai_total for c in comparisons]
        differences = [c.difference for c in comparisons]
        
        human_avg = statistics.mean(human_totals)
        ai_avg = statistics.mean(ai_totals)
        human_std = statistics.stdev(human_totals) if len(human_totals) > 1 else 0.0
        ai_std = statistics.stdev(ai_totals) if len(ai_totals) > 1 else 0.0
        
        avg_diff = statistics.mean(differences)
        abs_avg_diff = statistics.mean([abs(d) for d in differences])
        
        print(f"Sample Size: {len(comparisons)} submissions")
        print(f"")
        print(f"TOTAL POINTS COMPARISON:")
        print(f"  Human Average:     {human_avg:.2f} ± {human_std:.2f}")
        print(f"  AI Average:        {ai_avg:.2f} ± {ai_std:.2f}")
        print(f"  Average Difference: {avg_diff:.2f} (AI - Human)")
        print(f"  Avg Absolute Diff:  {abs_avg_diff:.2f}")
        
        if human_avg > 0:
            pct_diff = (avg_diff / human_avg) * 100
            print(f"  Percentage Diff:    {pct_diff:.1f}%")
        
        # Agreement statistics
        exact_matches = sum(1 for c in comparisons if abs(c.difference) < 0.01)
        close_matches = sum(1 for c in comparisons if abs(c.difference) <= 1.0)
        
        print(f"")
        print(f"AGREEMENT ANALYSIS:")
        print(f"  Exact Matches:      {exact_matches}/{len(comparisons)} ({exact_matches/len(comparisons)*100:.1f}%)")
        print(f"  Close Matches (±1): {close_matches}/{len(comparisons)} ({close_matches/len(comparisons)*100:.1f}%)")
        
        # Rubric-level summary
        if stats:
            avg_agreement = statistics.mean([s.agreement_rate for s in stats])
            print(f"  Avg Rubric Agreement: {avg_agreement:.1f}%")
            
            # Identify most problematic rubric items
            problematic = sorted(stats, key=lambda s: abs(s.ai_avg - s.human_avg), reverse=True)[:3]
            print(f"")
            print(f"MOST DIFFERENT RUBRIC ITEMS:")
            for i, stat in enumerate(problematic, 1):
                diff = stat.ai_avg - stat.human_avg
                desc = stat.description[:50] + "..." if len(stat.description) > 50 else stat.description
                print(f"  {i}. {desc}")
                print(f"     Human: {stat.human_avg:.2f}, AI: {stat.ai_avg:.2f}, Diff: {diff:.2f}")


async def evaluate_project(
    project_path: Path,
    backend_client: BackendClient,
    num_students: int,
    output_dir: Path
) -> None:
    """Evaluate a single project."""
    
    print(f"\nEvaluating project: {project_path.name}")
    
    grades_dir = project_path / "grades"
    assignments_dir = project_path / "assignments"
    
    if not grades_dir.exists() or not assignments_dir.exists():
        print(f"Skipping {project_path.name}: missing grades or assignments directory")
        return
    
    # Get all CSV files
    csv_files = list(grades_dir.glob("*.csv"))
    if not csv_files:
        print(f"No CSV files found in {grades_dir}")
        return
    
    # Collect all evaluation data for unified report
    all_evaluation_data = []
    
    # Process each CSV file
    for csv_file in csv_files:
        print(f"  Processing {csv_file.name}")
        
        # Parse CSV
        parser = CSVParser()
        rubric_items, human_grades = parser.parse_rubric_from_csv(csv_file)
        
        print(f"    Found {len(rubric_items)} rubric items and {len(human_grades)} students")
        
        # Select students to evaluate (randomly sampled)
        all_student_ids = list(human_grades.keys())
        if len(all_student_ids) <= num_students:
            student_ids = all_student_ids
            print(f"    Using all {len(student_ids)} available students")
        else:
            student_ids = random.sample(all_student_ids, num_students)
            print(f"    Randomly selected {len(student_ids)} students from {len(all_student_ids)} available")
        
        # Collect backend results
        backend_results = {}
        
        for idx, student_id in enumerate(student_ids):
            print(f"    Evaluating student {idx + 1}/{len(student_ids)}: {student_id}")
            
            # Load student files
            submission_dir = assignments_dir / f"submission_{student_id}"
            if not submission_dir.exists():
                print(f"      Submission directory not found: {submission_dir}")
                continue
            
            # Read and filter source files (same logic as Chrome extension)
            source_files = {}
            for file_path in submission_dir.rglob("*"):
                if file_path.is_file():
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            relative_path = file_path.relative_to(submission_dir)
                            content = f.read()
                            
                            # Apply same filtering logic as Chrome extension
                            processed_content = CSVParser.process_file_content(str(relative_path), content)
                            if processed_content is not None:  # None means file was filtered out
                                source_files[str(relative_path)] = processed_content
                    except Exception as e:
                        print(f"      Error reading {file_path}: {e}")
            
            if not source_files:
                print(f"      No source files found for {student_id}")
                continue
            
            # Filter and prepare rubric items for backend
            filtered_rubric_items = filter_rubric_items_for_backend(rubric_items)
            backend_rubric_items = []
            for rubric_item in filtered_rubric_items:
                item_dict = {
                    "id": rubric_item.id,
                    "description": rubric_item.description,
                    "points": rubric_item.points,
                    "type": rubric_item.type
                }
                if rubric_item.options:
                    # Convert letter-based options to backend format
                    if rubric_item.type == "RADIO":
                        # Options are in format: {"Q": {"text": "...", "points": "..."}, "W": {...}}
                        # Backend expects: {"Q": "option text with credit indicator", "W": "option text with credit indicator"}
                        item_dict["options"] = add_credit_indicators_to_radio_options(rubric_item.options)
                    else:
                        item_dict["options"] = rubric_item.options
                backend_rubric_items.append(item_dict)
            
            # Send to backend
            assignment_context = {
                "course_id": project_path.name,
                "assignment_id": project_path.name,  # Use project directory name for rubric mapping
                "submission_id": student_id,
                "assignment_name": f"{project_path.name} - {csv_file.stem}"
            }
            
            # Print radio button JSON for debugging
            print_radio_button_json(backend_rubric_items, f"Student {student_id}: ")
            
            try:
                results = await backend_client.grade_submission(
                    assignment_context=assignment_context,
                    source_files=source_files,
                    rubric_items=backend_rubric_items
                )
                backend_results[student_id] = results
            except Exception as e:
                print(f"      Error from backend: {e}")
        
        # Filter human grades to only include evaluated students
        evaluated_human_grades = {sid: human_grades[sid] for sid in student_ids if sid in human_grades}
        
        # Store evaluation data for unified report
        evaluation_data = {
            'csv_name': csv_file.stem,
            'rubric_items': rubric_items,
            'human_grades': evaluated_human_grades,
            'backend_results': backend_results
        }
        all_evaluation_data.append(evaluation_data)
        
        print(f"    Completed evaluation for {csv_file.stem}")
    
    # Generate unified report for the entire project
    if all_evaluation_data:
        print(f"  Creating unified report for {project_path.name}")
        unified_report_generator = UnifiedExcelReportGenerator()
        unified_output_path = output_dir / f"{project_path.name}_UNIFIED_evaluation.xlsx"
        
        unified_report_generator.create_unified_report(
            project_name=project_path.name,
            all_evaluation_data=all_evaluation_data,
            output_path=unified_output_path
        )
        
        print(f"  ✅ Unified report saved to: {unified_output_path}")


async def evaluate_project_with_points_analysis(
    project_path: Path,
    backend_client: BackendClient,
    num_students: int,
    output_dir: Path
) -> None:
    """Evaluate a project with comprehensive points analysis."""
    
    print(f"\nEvaluating project with points analysis: {project_path.name}")
    
    grades_dir = project_path / "grades"
    assignments_dir = project_path / "assignments"
    
    if not grades_dir.exists() or not assignments_dir.exists():
        print(f"Skipping {project_path.name}: missing grades or assignments directory")
        return
    
    # Get all CSV files
    csv_files = list(grades_dir.glob("*.csv"))
    if not csv_files:
        print(f"No CSV files found in {grades_dir}")
        return
    
    # Collect all evaluation data
    all_comparisons = []
    all_stats = []
    
    # Process each CSV file
    for csv_file in csv_files:
        print(f"  Processing {csv_file.name}")
        
        # Parse CSV
        parser = CSVParser()
        rubric_items, human_grades = parser.parse_rubric_from_csv(csv_file)
        
        print(f"    Found {len(rubric_items)} rubric items and {len(human_grades)} students")
        
        # Select students to evaluate
        all_student_ids = list(human_grades.keys())
        if len(all_student_ids) <= num_students:
            student_ids = all_student_ids
        else:
            student_ids = random.sample(all_student_ids, num_students)
        
        # Collect AI grading results
        ai_results = {}
        
        for student_id in student_ids:
            human_grade = human_grades[student_id]
            print(f"      Grading submission {student_id} ({human_grade.name})")
            
            # Load submission files
            student_dir = assignments_dir / student_id
            if not student_dir.exists():
                print(f"        ⚠️  Directory not found: {student_dir}")
                continue
            
            source_files = {}
            for file_path in student_dir.rglob("*"):
                if file_path.is_file():
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                            # Process file content with same logic as Chrome extension
                            processed_content = parser.process_file_content(str(file_path.relative_to(student_dir)), content)
                            if processed_content is not None:
                                source_files[str(file_path.relative_to(student_dir))] = processed_content
                    except Exception as e:
                        print(f"        ⚠️  Error reading {file_path}: {e}")
            
            if not source_files:
                print(f"        ⚠️  No valid source files found")
                continue
            
                        # Get AI grading
            try:
                # Filter and convert rubric items to format expected by backend client
                filtered_rubric_items = filter_rubric_items_for_backend(rubric_items)
                backend_rubric_items = []
                for item in filtered_rubric_items:
                    backend_item = {
                        "id": item.id,
                        "description": item.description,
                        "points": item.points,
                        "type": item.type.upper()
                    }
                    if item.options:
                        if str(item.type).upper() == "RADIO":
                            backend_item["options"] = add_credit_indicators_to_radio_options(item.options)
                        else:
                            backend_item["options"] = item.options
                    backend_rubric_items.append(backend_item)
                
                # Print radio button JSON for debugging
                print_radio_button_json(backend_rubric_items, f"Points Analysis - Student {student_id}: ")
                
                # Call backend client
                results = await backend_client.grade_submission(
                    assignment_context={
                        "course_id": project_path.name,
                        "assignment_id": project_path.name,  # Use project directory name for rubric mapping
                        "submission_id": student_id,
                        "assignment_name": f"{project_path.name} - {csv_file.stem}"
                    },
                    source_files=source_files,
                    rubric_items=backend_rubric_items
                )
                
                # Convert results to decisions format
                ai_decisions = {}
                for result in results:
                    # Create a mock decision object that matches expected structure
                    decision = type('obj', (object,), {
                        'verdict': type('obj', (object,), {
                            'decision': type('obj', (object,), {'value': 'CHECK' if result.decision else 'CROSS'})() if result.rubric_id.startswith('checkbox') else None,
                            'selected_option': result.decision if result.rubric_id.startswith('radio') else None
                        })()
                    })()
                    ai_decisions[result.rubric_id] = decision
                
                ai_results[student_id] = AIGradingResult(
                    submission_id=student_id,
                    decisions=ai_decisions
                )
                
                print(f"        ✅ AI grading completed ({len(ai_decisions)} decisions)")
                
            except Exception as e:
                print(f"        ❌ AI grading failed: {e}")
                continue
        
        # Calculate points comparison for this CSV
        if ai_results:
            comparisons = PointsCalculator.compare_submissions(human_grades, ai_results, rubric_items)
            stats = PointsCalculator.calculate_rubric_stats(comparisons, rubric_items)
            
            all_comparisons.extend(comparisons)
            all_stats.extend(stats)
            
            # Print results for this CSV
            print(f"\n📊 RESULTS FOR {csv_file.name}")
            print(f"{'='*60}")
            
            PointsReporter.print_overall_summary(comparisons, stats)
            PointsReporter.print_rubric_stats(stats)
            PointsReporter.print_submission_comparison(comparisons, limit=5)
    
    # Print overall project summary
    if all_comparisons:
        print(f"\n🎯 OVERALL PROJECT SUMMARY: {project_path.name}")
        print(f"{'='*80}")
        
        # Combine stats by rubric ID (averaging across CSV files)
        combined_stats = {}
        for stat in all_stats:
            if stat.rubric_id not in combined_stats:
                combined_stats[stat.rubric_id] = []
            combined_stats[stat.rubric_id].append(stat)
        
        # Average the stats
        final_stats = []
        for rubric_id, stat_list in combined_stats.items():
            if stat_list:
                avg_stat = RubricItemStats(
                    rubric_id=rubric_id,
                    description=stat_list[0].description,
                    rubric_type=stat_list[0].rubric_type,
                    max_points=stat_list[0].max_points,
                    human_avg=statistics.mean([s.human_avg for s in stat_list]),
                    ai_avg=statistics.mean([s.ai_avg for s in stat_list]),
                    human_std=statistics.mean([s.human_std for s in stat_list]),
                    ai_std=statistics.mean([s.ai_std for s in stat_list]),
                    agreement_rate=statistics.mean([s.agreement_rate for s in stat_list]),
                    sample_size=sum([s.sample_size for s in stat_list])
                )
                final_stats.append(avg_stat)
        
        PointsReporter.print_overall_summary(all_comparisons, final_stats)
        PointsReporter.print_rubric_stats(final_stats)
        PointsReporter.print_submission_comparison(all_comparisons, limit=10)


async def collect_evaluation_tasks(
    projects_dir: Path,
    num_students: int
) -> List[EvaluationTask]:
    """Collect all evaluation tasks from all projects and assignments."""
    tasks = []
    parser = CSVParser()
    
    project_dirs = [d for d in projects_dir.iterdir() if d.is_dir()]
    print(f"🔍 Collecting evaluation tasks from {len(project_dirs)} projects...")
    
    for project_dir in project_dirs:
        grades_dir = project_dir / "grades"
        assignments_dir = project_dir / "assignments"
        
        if not grades_dir.exists() or not assignments_dir.exists():
            print(f"⚠️  Skipping {project_dir.name}: missing grades or assignments directory")
            continue
            
        csv_files = list(grades_dir.glob("*.csv"))
        if not csv_files:
            print(f"⚠️  No CSV files found in {grades_dir}")
            continue
            
        print(f"  📂 {project_dir.name}: {len(csv_files)} assignments")
        
        for csv_file in csv_files:
            # Parse CSV
            rubric_items, human_grades = parser.parse_rubric_from_csv(csv_file)
            
            # Select students to evaluate
            all_student_ids = list(human_grades.keys())
            if len(all_student_ids) <= num_students:
                student_ids = all_student_ids
            else:
                student_ids = random.sample(all_student_ids, num_students)
            
            print(f"    📄 {csv_file.name}: {len(student_ids)} students selected from {len(all_student_ids)} available")
            
            for student_id in student_ids:
                # Load student files
                submission_dir = assignments_dir / f"submission_{student_id}"
                if not submission_dir.exists():
                    print(f"      ⚠️  Directory not found: {submission_dir}")
                    continue
                
                # Read and filter source files
                source_files = {}
                for file_path in submission_dir.rglob("*"):
                    if file_path.is_file():
                        try:
                            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                relative_path = file_path.relative_to(submission_dir)
                                content = f.read()
                                processed_content = CSVParser.process_file_content(str(relative_path), content)
                                if processed_content is not None:
                                    source_files[str(relative_path)] = processed_content
                        except Exception as e:
                            print(f"      ⚠️  Error reading {file_path}: {e}")
                
                if not source_files:
                    print(f"      ⚠️  No source files found for {student_id}")
                    continue
                
                # Filter and prepare rubric items for backend
                filtered_rubric_items = filter_rubric_items_for_backend(rubric_items)
                backend_rubric_items = []
                for rubric_item in filtered_rubric_items:
                    item_dict = {
                        "id": rubric_item.id,
                        "description": rubric_item.description,
                        "points": rubric_item.points,
                        "type": rubric_item.type
                    }
                    if rubric_item.options:
                        if rubric_item.type == "RADIO":
                            item_dict["options"] = add_credit_indicators_to_radio_options(rubric_item.options)
                        else:
                            item_dict["options"] = rubric_item.options
                    backend_rubric_items.append(item_dict)
                
                # Create assignment context
                assignment_context = {
                    "course_id": project_dir.name,
                    "assignment_id": project_dir.name,  # Use project directory name for rubric mapping
                    "submission_id": student_id,
                    "assignment_name": f"{project_dir.name} - {csv_file.stem}"
                }
                
                # Create evaluation task
                task = EvaluationTask(
                    project_name=project_dir.name,
                    csv_name=csv_file.stem,
                    student_id=student_id,
                    assignment_context=assignment_context,
                    source_files=source_files,
                    rubric_items=backend_rubric_items,
                    human_grade=human_grades[student_id],
                    rubric_items_list=rubric_items
                )
                tasks.append(task)
    
    print(f"✅ Collected {len(tasks)} total evaluation tasks")
    return tasks


async def process_evaluation_batch(
    tasks: List[EvaluationTask],
    backend_client: BackendClient,
    batch_num: int,
    total_batches: int
) -> List[Tuple[EvaluationTask, Any]]:
    """Process a batch of evaluation tasks concurrently."""
    batch_size = len(tasks)
    print(f"📦 Student Batch {batch_num}/{total_batches}: processing {batch_size} students...")
    
    async def evaluate_single_task(task: EvaluationTask) -> Tuple[EvaluationTask, Any]:
        """Evaluate a single task."""
        try:
            # Print radio button JSON for debugging
            print_radio_button_json(task.rubric_items, f"Concurrent - Student {task.student_id}: ")
            
            results = await backend_client.grade_submission(
                assignment_context=task.assignment_context,
                source_files=task.source_files,
                rubric_items=task.rubric_items
            )
            print(f"✅ Student {task.student_id} ({task.project_name}/{task.csv_name}) completed")
            return (task, results)
        except Exception as e:
            print(f"❌ Student {task.student_id} ({task.project_name}/{task.csv_name}) failed: {type(e).__name__}: {e}")
            return (task, None)
    
    # Process all tasks in the batch concurrently
    batch_start_time = asyncio.get_event_loop().time()
    results = await asyncio.gather(*[evaluate_single_task(task) for task in tasks])
    batch_duration = asyncio.get_event_loop().time() - batch_start_time
    
    successful_results = [r for r in results if r[1] is not None]
    failed_count = batch_size - len(successful_results)
    
    if failed_count > 0:
        print(f"⚠️  Student Batch {batch_num} completed in {batch_duration:.1f}s: {len(successful_results)}/{batch_size} successful, {failed_count} failed")
    else:
        print(f"✅ Student Batch {batch_num} completed in {batch_duration:.1f}s: All {batch_size} students successful")
    
    return results


async def evaluate_projects_concurrent(
    projects_dir: Path,
    backend_client: BackendClient,
    num_students: int,
    output_dir: Path,
    concurrent_batch_size: int = DEFAULT_CONCURRENT_BATCH_SIZE,
    points_analysis: bool = False
) -> None:
    """Evaluate multiple projects using concurrent batch processing."""
    
    # Step 1: Collect all evaluation tasks
    all_tasks = await collect_evaluation_tasks(projects_dir, num_students)
    
    if not all_tasks:
        print("❌ No evaluation tasks found")
        return
    
    # Step 2: Process tasks in concurrent batches
    total_batches = (len(all_tasks) + concurrent_batch_size - 1) // concurrent_batch_size
    print(f"🎯 Starting concurrent evaluation: {len(all_tasks)} students in {total_batches} batches (size: {concurrent_batch_size})")
    
    all_results = []
    overall_start_time = asyncio.get_event_loop().time()
    
    for batch_num in range(total_batches):
        start_idx = batch_num * concurrent_batch_size
        end_idx = min(start_idx + concurrent_batch_size, len(all_tasks))
        batch_tasks = all_tasks[start_idx:end_idx]
        
        batch_results = await process_evaluation_batch(
            batch_tasks, backend_client, batch_num + 1, total_batches
        )
        all_results.extend(batch_results)
    
    overall_duration = asyncio.get_event_loop().time() - overall_start_time
    successful_count = len([r for r in all_results if r[1] is not None])
    
    failed_count = len(all_tasks) - successful_count
    if failed_count > 0:
        print(f"\n⚠️  Concurrent evaluation completed in {overall_duration:.1f}s: {successful_count}/{len(all_tasks)} successful, {failed_count} failed")
    else:
        print(f"\n✅ Concurrent evaluation completed in {overall_duration:.1f}s: All {len(all_tasks)} students successful")
    print(f"   Throughput: {len(all_tasks)/overall_duration:.1f} students/second")
    
    # Step 3: Group results by project and generate reports
    await generate_reports_from_results(all_results, output_dir, points_analysis)


async def generate_reports_from_results(
    all_results: List[Tuple[EvaluationTask, Any]],
    output_dir: Path,
    points_analysis: bool = False
) -> None:
    """Generate reports from the batched evaluation results."""
    
    # Group results by project
    projects_data = {}
    
    for task, results in all_results:
        if results is None:
            continue
            
        project_name = task.project_name
        csv_name = task.csv_name
        
        if project_name not in projects_data:
            projects_data[project_name] = {}
        
        if csv_name not in projects_data[project_name]:
            projects_data[project_name][csv_name] = {
                'rubric_items': task.rubric_items_list,
                'human_grades': {},
                'backend_results': {}
            }
        
        projects_data[project_name][csv_name]['human_grades'][task.student_id] = task.human_grade
        projects_data[project_name][csv_name]['backend_results'][task.student_id] = results
    
    # Generate reports for each project
    print(f"\n📊 Generating reports for {len(projects_data)} projects...")
    
    for project_name, project_data in projects_data.items():
        print(f"  📈 Generating report for {project_name}")
        
        # Prepare evaluation data for unified report
        all_evaluation_data = []
        all_comparisons = []
        all_stats = []
        
        for csv_name, csv_data in project_data.items():
            evaluation_data = {
                'csv_name': csv_name,
                'rubric_items': csv_data['rubric_items'],
                'human_grades': csv_data['human_grades'],
                'backend_results': csv_data['backend_results']
            }
            all_evaluation_data.append(evaluation_data)
            
            if points_analysis:
                # Convert backend results to AIGradingResult format for points analysis
                ai_results = {}
                for student_id, results in csv_data['backend_results'].items():
                    ai_decisions = {}
                    for result in results:
                        decision = type('obj', (object,), {
                            'verdict': type('obj', (object,), {
                                'decision': type('obj', (object,), {'value': 'CHECK' if result.decision else 'CROSS'})() if result.rubric_id.startswith('checkbox') else None,
                                'selected_option': result.decision if result.rubric_id.startswith('radio') else None
                            })()
                        })()
                        ai_decisions[result.rubric_id] = decision
                    
                    ai_results[student_id] = AIGradingResult(
                        submission_id=student_id,
                        decisions=ai_decisions
                    )
                
                # Calculate points comparison
                if ai_results:
                    comparisons = PointsCalculator.compare_submissions(
                        csv_data['human_grades'], ai_results, csv_data['rubric_items']
                    )
                    all_comparisons.extend(comparisons)
                    
                    # Calculate stats
                    stats = PointsCalculator.calculate_rubric_stats(
                        csv_data['human_grades'], ai_results, csv_data['rubric_items']
                    )
                    all_stats.extend(stats)
        
        # Generate unified report
        if all_evaluation_data:
            unified_report_generator = UnifiedExcelReportGenerator()
            unified_output_path = output_dir / f"{project_name}_UNIFIED_evaluation.xlsx"
            
            unified_report_generator.create_unified_report(
                project_name=project_name,
                all_evaluation_data=all_evaluation_data,
                output_path=unified_output_path
            )
            
            print(f"    ✅ Unified report created: {unified_output_path}")
        
        # Print points analysis if enabled
        if points_analysis and all_comparisons:
            print(f"\n🎯 POINTS ANALYSIS: {project_name}")
            print(f"{'='*60}")
            PointsReporter.print_overall_summary(all_comparisons, all_stats)
            PointsReporter.print_rubric_stats(all_stats)
    
    # Generate cross-project unified report
    if projects_data:
        print(f"\n📊 Generating Cross-Project Unified Report...")
        try:
            cross_project_generator = CrossProjectUnifiedReportGenerator()
            cross_project_output_path = output_dir / "CROSS_PROJECT_UNIFIED_evaluation.xlsx"
            
            # Transform projects_data structure to match what CrossProjectUnifiedReportGenerator expects
            print("   🔄 Transforming data structure...")
            transformed_projects_data = {}
            for project_name, project_data in projects_data.items():
                transformed_projects_data[project_name] = []
                for csv_name, csv_data in project_data.items():
                    evaluation_data = {
                        'csv_name': csv_name,
                        'rubric_items': csv_data['rubric_items'],
                        'human_grades': csv_data['human_grades'],
                        'backend_results': csv_data['backend_results']
                    }
                    transformed_projects_data[project_name].append(evaluation_data)
            print("   ✅ Data structure transformation completed")
            
            # Add progress tracking
            total_projects = len(transformed_projects_data)
            total_assignments = sum(len(data_list) for data_list in transformed_projects_data.values())
            print(f"   Processing {total_projects} projects with {total_assignments} total assignments...")
            
            cross_project_generator.create_cross_project_report(
                all_project_data=transformed_projects_data,
                output_path=cross_project_output_path
            )
            
            print(f"✅ Cross-Project Unified Report created: {cross_project_output_path}")
            print(f"   📋 Report includes {len(transformed_projects_data)} projects with comprehensive statistics and analysis")
            
        except Exception as e:
            print(f"❌ Failed to generate cross-project unified report: {type(e).__name__}: {e}")
            print(f"   📍 Error occurred during report generation")
            import traceback
            traceback.print_exc()


async def main():
    """Main function to run evaluations."""
    parser = argparse.ArgumentParser(description="Evaluation Dashboard for Backend Grading")
    parser.add_argument("--projects-dir", type=str, default="evaluation_projects", 
                        help="Directory containing project folders")
    parser.add_argument("--backend-url", type=str, default="http://localhost:8000",
                        help="Backend API URL")
    parser.add_argument("--num-students", type=int, default=5,
                        help="Number of students to evaluate per CSV")
    parser.add_argument("--output-dir", type=str, default="evaluation_results",
                        help="Output directory for results")
    parser.add_argument("--points-analysis", action="store_true",
                        help="Enable comprehensive points comparison analysis")
    parser.add_argument("--concurrent-batch-size", type=int, default=DEFAULT_CONCURRENT_BATCH_SIZE,
                        help="Number of students to process concurrently (default: 20)")
    parser.add_argument("--use-concurrent", action="store_true", default=True,
                        help="Use concurrent batch processing for improved performance")
    parser.add_argument("--sequential", action="store_true",
                        help="Use sequential processing (disables concurrent processing)")
    
    args = parser.parse_args()
    
    # Setup
    projects_dir = Path(args.projects_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)
    
    if not projects_dir.exists():
        print(f"Projects directory not found: {projects_dir}")
        return
    
    # Create backend client
    backend_client = BackendClient(args.backend_url)
    
    # Test backend connection
    try:
        health_status = await backend_client.check_health()
        print(f"✅ Backend connection successful: {health_status}")
    except Exception as e:
        print(f"❌ Backend connection failed: {e}")
        print("Make sure the backend server is running!")
        return
    
    # Find all project directories
    project_dirs = [d for d in projects_dir.iterdir() if d.is_dir()]
    if not project_dirs:
        print(f"No project directories found in {projects_dir}")
        return
    
    # Determine processing mode
    use_concurrent = args.use_concurrent and not args.sequential
    
    print(f"Found {len(project_dirs)} project(s) to evaluate")
    print(f"Backend URL: {args.backend_url}")
    print(f"Students per CSV: {args.num_students}")
    print(f"Output directory: {output_dir}")
    print(f"Points analysis: {'Enabled' if args.points_analysis else 'Disabled'}")
    print(f"Processing mode: {'Concurrent' if use_concurrent else 'Sequential'}")
    if use_concurrent:
        print(f"Concurrent batch size: {args.concurrent_batch_size}")
    
    # Process projects using chosen method
    try:
        if use_concurrent:
            # Use new concurrent batch processing
            await evaluate_projects_concurrent(
                projects_dir, backend_client, args.num_students, output_dir,
                args.concurrent_batch_size, args.points_analysis
            )
        else:
            # Use original sequential processing
            for project_dir in project_dirs:
                try:
                    if args.points_analysis:
                        await evaluate_project_with_points_analysis(
                            project_dir, backend_client, args.num_students, output_dir
                        )
                    else:
                        await evaluate_project(
                            project_dir, backend_client, args.num_students, output_dir
                        )
                except Exception as e:
                    print(f"❌ Error processing {project_dir.name}: {e}")
                    continue
    except Exception as e:
        print(f"❌ Error during evaluation: {e}")
        return
    
    print(f"\n🎉 Evaluation complete! Results saved to: {output_dir}")


if __name__ == "__main__":
    asyncio.run(main()) 